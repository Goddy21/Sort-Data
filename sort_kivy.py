from kivy.config import Config
Config.set('graphics', 'resizable', False)
Config.set('graphics', 'width', '600')
Config.set('graphics', 'height', '350')
Config.set('graphics', 'position', 'custom')
Config.set('graphics', 'left', '100')
Config.set('graphics', 'top', '100')

import os
import pandas as pd
import datetime
import re

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.utils import get_color_from_hex
from kivy.uix.progressbar import ProgressBar
from kivy.clock import Clock
from kivy.graphics import Color, Rectangle
from plyer import filechooser

Window.clearcolor = get_color_from_hex('#F0F4F8')
Window.size = (600, 350)
Window.resizable = False

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
sort_result_folder = os.path.join(desktop_path, "SORT RESULT")
os.makedirs(sort_result_folder, exist_ok=True)

class ExcelFilterApp(App):
    def build(self):
        self.icon = "icon.png"
        self.title = "Mkononi Numbers-Sorting System"

        self.root_layout = BoxLayout(orientation='vertical', padding=20, spacing=10)

        banner = Label(
            text="Filter and sort phone numbers based on names",
            size_hint_y=None,
            height=50,
            color=get_color_from_hex('#FFFFFF'),
            bold=True,
            font_size=20
        )
        with banner.canvas.before:
            Color(*get_color_from_hex('#0056b3'))
            self.rect = Rectangle(size=banner.size, pos=banner.pos)
        banner.bind(size=self._update_rect, pos=self._update_rect)
        self.root_layout.add_widget(banner)

        input_grid = GridLayout(cols=2, spacing=15, size_hint_y=0.5)

        input_grid.add_widget(Label(text="[b]Input folder path:[/b]", markup=True, color=get_color_from_hex('#333333'), font_size=15))
        folder_layout = BoxLayout(orientation='horizontal')
        self.folder_var = TextInput(multiline=False)
        browse_button = Button(text="Browse", background_color=get_color_from_hex('#007BFF'))
        browse_button.bind(on_release=self.show_file_chooser)
        folder_layout.add_widget(self.folder_var)
        folder_layout.add_widget(browse_button)
        input_grid.add_widget(folder_layout)

        input_grid.add_widget(Label(text="[b]Name(s) to search:[/b]", markup=True, color=get_color_from_hex('#333333'), font_size=15))
        self.search_name_var = TextInput(multiline=True)
        input_grid.add_widget(self.search_name_var)

        self.root_layout.add_widget(input_grid)

        self.progress_bar = ProgressBar(max=100, value=0, size_hint_y=0.05)
        self.root_layout.add_widget(self.progress_bar)

        self.execute_button = Button(text="Execute Filter", size_hint_y=0.1, background_color=get_color_from_hex('#28A745'))
        self.execute_button.bind(on_release=self.start_filtering)
        self.root_layout.add_widget(self.execute_button)

        return self.root_layout

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def show_file_chooser(self, instance):
        try:
            os.environ['PLYER_DEFAULT_PROVIDER'] = 'win32'
            filechooser.choose_dir(title="Select folder", on_selection=self.selected)
        except Exception as e:
            self.show_popup("Error", str(e))

    def selected(self, filename):
        if filename:
            self.folder_var.text = filename[0]

    def start_filtering(self, instance):
        self.progress_bar.value = 0
        self.execute_button.disabled = True

        self.folder = self.folder_var.text
        raw_names = self.search_name_var.text.strip().lower()
        self.search_names = list({name.strip() for name in re.split(r'[\n\r,]+', raw_names) if name.strip()})

        if not os.path.isdir(self.folder):
            self.show_popup("Input Error", "Invalid folder path. Please select a valid directory.")
            self.execute_button.disabled = False
            return

        if not self.search_names:
            self.show_popup("Input Error", "Please enter at least one name.")
            self.execute_button.disabled = False
            return

        self.files_created = 0
        self.required_columns = ['Credit Identity String', 'Customer Name']
        self.total = len(self.search_names)
        self.index = 0
        self.all_matches = pd.DataFrame(columns=self.required_columns)
        self.log_summary = []

        Clock.schedule_once(self.process_next_name, 0.1)

    def process_next_name(self, dt):
        if self.index >= self.total:
            self.execute_button.disabled = False
            
            if "Credit Identity String" in self.all_matches.columns:
                real_matches = self.all_matches[
                    ~self.all_matches["Credit Identity String"].astype(str).str.startswith("-----")
                ]
            else:
                real_matches = pd.DataFrame()


            if not real_matches.empty:
                # Save the combined results
                self.all_matches.drop_duplicates(subset='Credit Identity String', inplace=True)
                combined_file = os.path.join(sort_result_folder, f"combined_matches_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                self.all_matches.to_excel(combined_file, index=False)

                # Apply formatting
                from openpyxl import load_workbook
                from openpyxl.styles import Font
                wb = load_workbook(combined_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    cell_value = row[0].value
                    if isinstance(cell_value, str) and cell_value.startswith("----- RESULTS FOR"):
                        row[0].font = Font(bold=True, size=14, color="FF0000")
                wb.save(combined_file)

                message = f"Found {len(real_matches)} matching row(s)."
            else:
                message = "No matches found."

            # Save log file if needed
            if self.log_summary:
                log_file = os.path.join(sort_result_folder, f"log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                with open(log_file, 'w') as f:
                    f.write("\n".join(self.log_summary))

            self.show_popup("Done", message)
            return


        search_name = self.search_names[self.index]
        filtered_data = pd.DataFrame(columns=self.required_columns)

        print(f"Searching for: '{search_name}'")  # Debugging line

        # Start processing files
        for filename in os.listdir(self.folder):
            filepath = os.path.join(self.folder, filename)

            try:
                print(f"Processing file: {filename}")  # Debugging line

                # Load Excel files
                if filename.endswith(('.xlsx', '.xlsm')):
                    xls = pd.ExcelFile(filepath)

                    for sheet_name in xls.sheet_names:
                        df = xls.parse(sheet_name)
                        df.columns = df.columns.str.strip()  # Clean up column names

                        print(f"Found columns: {df.columns.tolist()}")  # Debugging

                        # Check the expected columns
                        if 'Customer Name' in df.columns and 'Credit Identity String' in df.columns:
                            # Display customer name samples for debugging
                            print(f"Customer Names sample: {df['Customer Name'].head().tolist()}")  # Print first few Customer Names

                            # Perform search
                            rows = df[df['Customer Name'].str.contains(rf'\b{re.escape(search_name)}\b', case=False, na=False, regex=True)]
                            filtered_data = pd.concat([filtered_data, rows[self.required_columns]], ignore_index=True)

                            # Debugging statement to show how many matches were found
                            print(f"Matches found for '{search_name}' in {filename}: {len(rows)}")  # Debugging

                # Load CSV files
                elif filename.endswith('.csv'):
                    df = pd.read_csv(filepath, on_bad_lines='skip')
                    df.columns = df.columns.str.strip()  # Clean up column names

                    if 'Customer Name' in df.columns and 'Credit Identity String' in df.columns:
                        rows = df[df['Customer Name'].str.contains(rf'\b{re.escape(search_name)}\b', case=False, na=False, regex=True)]
                        filtered_data = pd.concat([filtered_data, rows[self.required_columns]], ignore_index=True)

                        print(f"Matches found for '{search_name}' in {filename}: {len(rows)}")  # Debugging

            except Exception as e:
                print(f"[!] Error reading {filename}: {e}")

        if not filtered_data.empty:
            print(f"Filtered data for '{search_name}':")  # Debugging line
            print(filtered_data)  # Debugging line

            # Append matched data
            self.all_matches = pd.concat([self.all_matches, filtered_data], ignore_index=True)
        else:
            print(f"No filtered data found for: '{search_name}'")  # Debugging to verify

        # Add separator regardless of whether there was data or not
        separator = pd.DataFrame([{
            "Credit Identity String": f"----- RESULTS FOR: {self.search_names[self.index].upper()} (NO MATCHES FOUND)",
            "Customer Name": ""
        }]) if filtered_data.empty else pd.DataFrame([{
            "Credit Identity String": f"----- RESULTS FOR: {self.search_names[self.index].upper()}",
            "Customer Name": ""
        }])

        self.all_matches = pd.concat([self.all_matches, separator], ignore_index=True)

        # Move to the next name
        self.index += 1
        self.progress_bar.value = (self.index / self.total) * 100

        # Continue to the next name
        Clock.schedule_once(self.process_next_name, 0.1)


    def show_popup(self, title, message):
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        layout.add_widget(Label(text=message, halign='center', valign='middle', text_size=(380, None)))
        btn = Button(text="Close", size_hint=(1, 0.3), background_color=get_color_from_hex('#007BFF'), color=(1, 1, 1, 1))
        popup = Popup(title=title, content=layout, size_hint=(None, None), size=(400, 200), auto_dismiss=False)
        btn.bind(on_release=popup.dismiss)
        layout.add_widget(btn)
        popup.open()

if __name__ == '__main__':
    ExcelFilterApp().run()
